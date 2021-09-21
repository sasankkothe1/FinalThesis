import os
from pathlib import Path
from datetime import date, datetime
from flask import Flask, json, request, jsonify
from flask.globals import current_app
from flask_cors import CORS
from flask.helpers import send_from_directory,  send_file
from flask_mail import Mail, Message

from pymongo import MongoClient
from openpyxl import Workbook
from bson.json_util import dumps

# for creating secret token for the user authentication
from flask_jwt_extended import create_access_token, get_jwt_identity, jwt_required, JWTManager


from extractSolutionFiles import extractSolutionFiles

app = Flask(__name__, static_folder='../frontend/build', static_url_path='/')
CORS(app, resources={r"/*": {"origins": "*"}})

# mail configs
app.config['MAIL_SERVER'] = 'outlook.office365.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USERNAME'] = 'pspliboperationsmanagement@outlook.com'
app.config['MAIL_PASSWORD'] = 'psplib123$'
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_ASCII_ATTACHMENTS'] = True
app.config['MAIL_DEBUG'] = True
app.config['MAIL_SUPPRESS_SEND'] = False
app.config['GENERATED_REPORT_FOLDER'] = './'
app.config['DOWNLOAD_FOLDER'] = '../../fileStore/problemSets'

# adding the secret key for the JWT token
app.config["JWT_SECRET_KEY"] = "secretsuperpsplibthesissasanktum"
jwt = JWTManager(app)

mail = Mail(app)

# client = MongoClient(
#     "mongodb+srv://testpsplib:testpsplib@cluster0.trtwg.mongodb.net/myFirstDatabase?retryWrites=true&w=majority")
# client = MongoClient(
#     "mongodb://localhost:27017/?readPreference=primary&appname=MongoDB%20Compass&directConnection=true&ssl=false")

client = MongoClient("mongodb://localhost:27017/")
db = client['psplib2']
print(db)

ProblemSetsLocation = "../../fileStore/problemSets"
uploadedFilesLocation = "../../fileStore/UploadedSolutions"
ubBestSolutionsLocation = "../../fileStore/bestSolutions/rcpsp/hrs"
lbBestSolutionsLocation = "../../fileStore/bestSolutions/rcpsp/lb"
rcpspOPTBestSolutionsLocation = "../../fileStore/bestSolutions/rcpsp/opt"


@app.route('/')
def index():
    return app.send_static_file('index.html')


@app.route('/login', methods=['POST'])
def login():
    username = request.form['username']
    password = request.form['password']

    if username == "psplib-admin" and password == "psplib-admin-password":
        return "/Admin", 200

    elif username != "psplib-secret-username" or password != "psplib-secret-password":
        return "wrong username and password", 404
    accessToken = create_access_token(identity=username)
    return jsonify(accessToken=accessToken)


@app.route('/getSubmissions', methods=['GET'])
def getSubmissions():
    submissionCollection = db['submissions']
    result = submissionCollection.find()
    return dumps(list(result))


@app.route('/getBestResults', methods=['GET'])
def getBestResults():
    bestResultsCollection = db['bestresults']
    result = bestResultsCollection.find()
    return dumps(list(result))


@app.route('/getJobOptions')
def getJobOptions():
    submissionCollection = db['submissions']
    jobNumbers = list(submissionCollection.find({}, {"jobs": 1, "_id": 0}))
    jobNumbersList = []

    for jobNumber in jobNumbers:
        if {"value": (jobNumber["jobs"]+"").rstrip(), "label": (jobNumber["jobs"]+"").rstrip()} not in jobNumbersList:
            jobNumbersList.append(
                {"value": (jobNumber["jobs"]+"").rstrip(), "label": (jobNumber["jobs"]+"").rstrip()})
    return dumps(jobNumbersList)


@app.route('/getReport', methods=['GET'])
def getReport():
    try:
        jobs = request.args.get('jobs')
        mode = request.args.get('mode')
        type = request.args.get('type')
        reportType = request.args.get('reportType')
        startDate = request.args.get('startDate')
        endDate = request.args.get('endDate')
        submissionCollection = db['submissions']
        bestResultsCollection = db['bestresults']

        if reportType == "reportFiles":
            if startDate and endDate:
                startDateArray = startDate.split("-")
                endDateArray = endDate.split("-")
                fromDate = datetime(day=int(startDateArray[0]), month=int(
                    startDateArray[1]), year=int(startDateArray[2]))
                toDate = datetime(day=int(endDateArray[0]), month=int(
                    endDateArray[1]), year=int(endDateArray[2]))
                reportData = submissionCollection.find({"mode": mode, "jobs": jobs, "submissionDate": {
                                                       "$gte": fromDate, "$lte": toDate}}, {"_id": 0})
                reportName = generateReport(reportData, type, reportType)
                return send_from_directory(app.config['GENERATED_REPORT_FOLDER'], filename=reportName, as_attachment=True)
            else:
                reportData = submissionCollection.find(
                    {"mode": mode, "jobs": jobs}, {"_id": 0})
                reportName = generateReport(reportData, type, reportType)
                return send_from_directory(app.config['GENERATED_REPORT_FOLDER'], filename=reportName, as_attachment=True)
        if reportType == "summaryFiles":
            reportData = bestResultsCollection.find(
                {"mode": mode, "jobs": jobs}, {"_id": 0})
            reportName = generateReport(reportData, type, reportType)
            return send_from_directory(app.config['GENERATED_REPORT_FOLDER'], filename=reportName, as_attachment=True)

    except Exception as e:
        print("exception is ", e)
    finally:
        os.remove(reportName)


def generateReport(reportData, type, reportType):
    # TODO: implement excel in the morning
    filename = "report_" + \
        str(type) + "_" + str(datetime.today().strftime("%d_%m_%Y")) + ".xlsx"

    if type == 'hrs':
        if reportType == "reportFiles":
            columnHeaders = ['Par', 'Inst', 'Makespan',
                             'Submitted At', 'Submitted By']
            keys = ['par', 'inst', 'ub', 'submissionDate', 'name']
        if reportType == "summaryFiles":
            columnHeaders = ['Par', 'Inst', 'Best UB',
                             'Submitted By', 'Submitted At']
            keys = ['par', 'inst', 'ub', 'AuthorUB', 'submissionDateUB']
    if type == 'lb':
        if reportType == "reportFiles":
            columnHeaders = ['Par', 'Inst', 'Lower Bound',
                             'Submitted At', 'Submitted By']
            keys = ['par', 'inst', 'lb', 'submissionDate', 'name']
        if reportType == "summaryFiles":
            columnHeaders = ['Par', 'Inst', 'Best Lower Bound',
                             'Submitted By', 'Submitted At']
            keys = ['par', 'inst', 'lb', 'AuthorLB', 'submissionDateLB']

    return excelSheetGenerator(reportData, columnHeaders, keys, filename)


def excelSheetGenerator(reportData, columnHeaders, keys, filename):
    workbook = Workbook()
    sheet = workbook.active
    row = 1
    col = 1
    for header in columnHeaders:
        sheet.cell(row=row, column=col, value=header)
        col += 1
    col = 1
    for el in reportData:
        row += 1
        for key in keys:
            sheet.cell(row, col, value=el[key])
            col += 1
        col = 1

    workbook.save(filename)
    return filename


@ app.route('/getTheFileList', methods=['GET'])
def getTheFileList():
    problemType = request.args.get('problemType')
    mode = request.args.get('mode')
    pathToDirectory = ProblemSetsLocation + "/" + problemType + "/" + mode
    submissionCollection = db['submissions']

    return jsonify(os.listdir(pathToDirectory))


@ app.route('/downloadFile', methods=['GET'])
def downloadFile():
    problemType = request.args.get('problemType')
    mode = request.args.get('mode')
    fileName = request.args.get('fileName')
    pathToDirectory = app.config['DOWNLOAD_FOLDER'] + \
        "/" + problemType + "/" + mode
    finalDirectory = os.path.join(current_app.root_path, pathToDirectory)
    try:
        return send_from_directory(finalDirectory, fileName, as_attachment=True, attachment_filename=fileName)
    except FileNotFoundError:
        print("fileNotFound")


@ app.route('/upload', methods=['POST'])
def uploadFile():
    fileArray = request.files.getlist('files')
    name = request.form['name']
    uploadTime = str(datetime.now())
    typeOfInstance = request.form['typeOfInstance']

    deniedFileExtensions = ['mp4', 'mp3', 'py', 'java',
                            'class', 'php', 'png', 'jpg', 'jpeg', 'gif']

    pathToBeStoredAt = uploadedFilesLocation + "/" + name + "/" + uploadTime

    for file in fileArray:
        fileExtension = file.filename.split('.')[-1].lower()
        if fileExtension in deniedFileExtensions:
            emailID = request.form['email']
            msg = Message("Error in the files Uploaded",
                          sender="pspliboperationsmanagement@outlook.com", recipients=[emailID])
            msg.body = "You have uploaded files with wrong extensions. Please check" + \
                file.filename+" before uploading"
            mail.send(msg)
            return "your files have unsupported format. Please upload only text, pdf, or zip files", 422
        else:
            Path(pathToBeStoredAt).mkdir(
                mode=0o777, parents=True, exist_ok=True)
            file.save(os.path.join(pathToBeStoredAt, file.filename))

    userData = {'name': request.form['name'], 'email': request.form['email'], 'titleOfPaper': request.form['titleOfPaper'],
                'contributors': request.form['contributors'], 'typeOfInstance': typeOfInstance, 'typeOfSolution': request.form['typeOfSolution']}

    processSolution(userData, pathToBeStoredAt + "/")
    return 'upload successful', 200


# ImmutableMultiDict([('name', 'Sasank Kothe'), ('email', 'sasankkothe@gmail.com'),
# ('titleOfPaper', ''), ('contributors', ''), ('typeOfInstance', 'rcpsp_sm')])

def processSolution(userData, solutionFilesPath):
    # extracting individual data from dictionary
    typeOfInstance = userData['typeOfInstance']
    typeOfSolution = userData['typeOfSolution']

    answer = extractSolutionFiles(
        solutionFilesPath, typeOfInstance, typeOfSolution)

    # store the submission into the MondoDB client collection = submissions
    storeSubmission(userData, answer)

    # to check if every solution is best or not and create report
    reportList = []

    for el in answer:
        reportList.append(isSubmissionBest(userData, el))

    # create the report.xlsx
    reportCreated = createReport(reportList)

    if reportCreated:
        sendEmail(userData)
        reportCreated = False
        reportList = []

    # convert the answer to json and send the answer
    return jsonify(answer)


def storeSubmission(userData, answer):

    # username = "psplib-test-user"
    # password = "psplib-test-password"

    # access_token = create_access_token(identity=username)

    name = userData['name']
    email = userData['email']
    titleOfPaper = userData['titleOfPaper']
    contributors = userData['contributors']
    typeOfInstance = userData['typeOfInstance']
    typeOfSolution = userData['typeOfSolution']

    submissionCollection = db['submissions']

    # ('e._goncharov,_v._leonov_16-07-2015_00-05-14_j12011_10.sm', 'j120', '11',
    # '10', '../../fileStore/UploadedSolutions/Sasank_Kothe/15_4_2021/18_41_24/', 180, False, None)
    for el in answer:
        post = {
            "name": name,
            "email": email,
            "titleOfPaper": titleOfPaper,
            "contributors": contributors,
            "typeOfInstance": typeOfInstance,
            "typeOfSolution": typeOfSolution,
            "submissionDate": datetime.now(),
            "mode": typeOfInstance.split("_")[1],
            "fileName": str(el[0]).rstrip(),
            "jobs": str(el[1]).rstrip(),
            "par": str(el[2]).rstrip(),
            "inst": str(el[3]).rstrip(),
            "ub": str(el[5]).rstrip(),
            "lb": str(el[6]).rstrip(),
            "fileLocation": str(el[4]).rstrip(),
            "isError": el[7],
            "error": el[8]
        }
        submissionCollection.insert_one(post)

# 'el' is a tuple of (0: instanceFileName, 1: jobNumber, 2: par number, 3: inst number, 4: locationOfTheFileStore, 5: ub, 6: isError, 7: error)


def isSubmissionBest(userData, el):

    typeOfSolution = userData['typeOfSolution']
    typeOfInstance = userData['typeOfInstance']

    # get the location where the file is originally stored in "originalFileLocation" variable
    fileName = el[0]
    originalLocationStoredDirectory = el[4]
    originalFileLocation = originalLocationStoredDirectory + fileName

    bestResultsCollection = db['bestresults']

    ubSwap = False
    lbSwap = False
    ubDeviationPercentage = 0
    lbDeviationPercentage = 0
    docUB = 0
    docLB = 0
    objectiveFunc = el[5]
    submittedLB = el[6]
    feasible = False

    instance = el[1]+el[2]+"_"+el[3]

    if el[7] == False:
        feasible = True

    query = {"jobs": el[1], "par": el[2], "inst": el[3],
             "instanceType": typeOfInstance}

    foundResult = bestResultsCollection.find(
        query)   # a mongo cursor is returned

    # check if particular solution is present in the bestSolutions table. if not, insert
    if foundResult.count() == 0:
        if typeOfSolution == "Upper Bound":

            if feasible:
                post = {
                    "instanceType": typeOfInstance,
                    "mode": typeOfInstance.split("_")[1],
                    "jobs": el[1],
                    "par": el[2],
                    "inst": el[3],
                    "ub": el[5],
                    "lb": el[6],
                    "submissionDateUB": datetime.now(),
                    "submissionDateLB": "not available",
                    "AuthorUB": userData['name'],
                    "AuthorLB": "not available"
                }
                bestResultsCollection.insert_one(post)

                # TODO: ask the deviationUB in the below line (3rd element from right)
                return ((instance, userData['typeOfInstance'], feasible, objectiveFunc, 0, objectiveFunc, 0, 0, True, False))
            else:
                return ((instance, userData['typeOfInstance'], feasible, objectiveFunc, 0, 0, 0, 0, ubSwap, lbSwap))

        if typeOfSolution == "Lower Bound":
            post = {
                "instanceType": userData['typeOfInstance'],
                "mode": typeOfInstance.split("_")[1],
                "jobs": el[1],
                "par": el[2],
                "inst": el[3],
                "ub": el[5],
                "lb": el[6],
                "submissionDateUB": "not available",
                "submissionDateLB": datetime.now(),
                "AuthorUB": "not available",
                "AuthorLB": userData['name']
            }
            bestResultsCollection.insert_one(post)
            return ((instance, userData['typeOfInstance'], feasible, 0, submittedLB, 0, 0, 0, True, False))
    # if the solution is already present in the best solution
    else:
        for result in foundResult:
            if (typeOfSolution == "Upper Bound"):
                # store the originalFileobject in "originalFile" variable
                # print(originalFileLocation)
                originalFile = open(originalFileLocation)

                docUB = result['ub']
                userUB = el[5]

                if int(userUB) < int(docUB):
                    bestResultsCollection.update(
                        query, {"$set": {"ub": userUB, "AuthorUB": userData['name'], "submissionDateUB": datetime.now()}})
                    ubSwap = True
                    ubDeviationPercentage = (
                        (int(userUB) - int(docUB)) / int(docUB)) * 100

                    # store the file in the best files directory --> UB
                    # check if the file is already present and remove it
                    if os.path.exists(ubBestSolutionsLocation + "/" + fileName):
                        os.remove(ubBestSolutionsLocation + "/" + fileName)

                    # save the best UB file in the hrs folder of the best solution
                    # create new file and copy the contents of the originalFile into the new file
                    with open(ubBestSolutionsLocation + "/" + fileName, 'a') as newBestfile:
                        for line in originalFile:
                            newBestfile.write(line)
                        newBestfile.close()

            if (typeOfSolution == "Lower Bound"):

                docLB = result['lb']
                userLB = el[6]

                if int(userLB) > int(docLB):
                    bestResultsCollection.update(
                        query, {"$set": {"lb": userLB, "AuthorLB": userData['name'], "submissionDateLB": datetime.now()}})
                lbSwap = True
                lbDeviationPercentage = (
                    (int(userLB) - int(docLB)) / int(docLB)) * 100

        return ((instance, userData['typeOfInstance'], feasible, objectiveFunc, docLB, docUB, lbDeviationPercentage, ubDeviationPercentage, ubSwap, lbSwap))


def createReport(reportList):

    # reportList: (instance, userData['typeOfInstance'], feasible, objectiveFunc, docLB, docUB, lbDeviationPercentage, ubDeviationPercentage, ubSwap, lbSwap)
    workbook = Workbook()
    sheet = workbook.active

    sheet.merge_cells('C1:D1')
    sheet.merge_cells('E1:F1')
    sheet.merge_cells('G1:I1')

    sheet['C1'] = "Submitted Solution"
    sheet['E1'] = "PSPLIB Solution"
    sheet['G1'] = "Results of Solution Check"

    row = 2
    col = 1
    columnHeaders = ['Instance', 'Problem Type', 'Feasible', 'Objective Function',
                     'CP-based LB', 'Best Known UB', 'Deviation from LB', 'Deviation from UB', 'Improved UB', 'Improved LB']

    for header in columnHeaders:
        sheet.cell(column=col, row=row, value=header)
        col += 1
    col = 1

    for el in reportList:
        row += 1
        for i in range(0, len(el)):
            sheet.cell(column=col, row=row, value=str(el[i]))
            col += 1
        col = 1

    workbook.save(filename="report.xlsx")

    return True


def sendEmail(userData):
    # application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
    msg = Message('Report from PSPLIB', sender="pspliboperationsmanagement@outlook.com",
                  recipients=[userData['email']])
    msg.body = "Please find the attachment as the status report of the submission"

    with app.open_resource("report.xlsx") as fp:
        msg.attach(
            "report.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", fp.read())

    mail.send(msg)

    os.remove("report.xlsx")


if __name__ == "__main__":
    # debug, when set to true, automatically reloads the server on file change and saved
    app.run(debug=True)
